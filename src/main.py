import os
import csv
import json
import subprocess
import tempfile
import shutil
import socket
import traceback
import uuid
import concurrent.futures
import asyncio
import threading
from pathlib import Path
from typing import List, Optional, Literal, Dict, Any
from datetime import timedelta
from enum import Enum
import boto3
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware

try:
    import openpyxl
except ImportError:
    openpyxl = None

load_dotenv()

app = FastAPI()

# In-memory task storage
tasks = {}
tasks_lock = threading.Lock()


class TaskStatus(str, Enum):
    IN_PROGRESS = "in_progress"
    CANCELLED = "cancelled"
    DONE = "done"


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class STTEvaluationRequest(BaseModel):
    audio_paths: List[str]  # S3 paths to audio files
    texts: List[str]  # Ground truth text for each audio file
    providers: List[
        str
    ]  # List of STT providers (e.g., ["deepgram", "openai", "sarvam"])
    language: str  # Language (e.g., "english", "hindi")


class ProviderResult(BaseModel):
    provider: str
    success: bool
    message: str
    metrics: Optional[List[Dict[str, Any]]] = None
    results: Optional[List[Dict[str, Any]]] = None


class TaskCreateResponse(BaseModel):
    task_id: str
    status: str


class TaskStatusResponse(BaseModel):
    task_id: str
    status: str
    provider_results: Optional[List[ProviderResult]] = None
    leaderboard_metrics_path: Optional[str] = None
    leaderboard_summary: Optional[List[Dict[str, Any]]] = None
    error: Optional[str] = None


class PresignedURLRequest(BaseModel):
    task_type: Literal["stt", "tts", "agent"]
    content_type: str  # e.g., "audio/wav", "text/csv"
    extension: str  # e.g., "wav", "csv"


class PresignedURLResponse(BaseModel):
    presigned_url: str
    s3_path: str
    expires_in: int  # expiration time in seconds


def is_port_in_use(port: int) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(("localhost", port)) == 0


def find_available_port(start_port: int = 8000) -> int:
    """Find an available port starting from start_port."""
    port = start_port
    while True:
        print(f"Checking port {port}")
        if is_port_in_use(port):
            port += 1
            if port > 65535:
                raise RuntimeError("No available ports found")
            continue

        return port


def get_s3_client():
    """Get S3 client from environment variables."""
    aws_access_key_id = os.getenv("AWS_ACCESS_KEY_ID")
    aws_secret_access_key = os.getenv("AWS_SECRET_ACCESS_KEY")
    aws_region = os.getenv("AWS_REGION", "ap-south-1")

    if aws_access_key_id and aws_secret_access_key:
        return boto3.client(
            "s3",
            aws_access_key_id=aws_access_key_id,
            aws_secret_access_key=aws_secret_access_key,
            region_name=aws_region,
        )

    return boto3.client("s3", region_name=aws_region)


def get_s3_output_config():
    """Get S3 output configuration from environment variables."""
    bucket = os.getenv("S3_OUTPUT_BUCKET")

    if not bucket:
        raise ValueError("S3_OUTPUT_BUCKET environment variable is required")

    return bucket


@app.get("/")
def read_root():
    return {"message": "Health check successful!"}


@app.post("/presigned-url", response_model=PresignedURLResponse)
async def get_presigned_url(request: PresignedURLRequest):
    """
    Generate a presigned URL for uploading files to S3.

    The file will be stored at: bucket/task_type/media/UUID.extension

    Args:
        request: Contains task_type (stt, tts, agent) and file_extension

    Returns:
        Presigned URL, S3 path, and expiration time
    """
    # Validate file extension (remove leading dot if present)
    file_extension = request.extension
    if not file_extension:
        raise HTTPException(
            status_code=400,
            detail="File extension cannot be empty",
        )

    # Validate task type
    if request.task_type not in ["stt", "tts", "agent"]:
        raise HTTPException(
            status_code=400,
            detail="task_type must be one of: stt, tts, agent",
        )

    # Get S3 bucket from environment
    s3_bucket = os.getenv("S3_OUTPUT_BUCKET")
    if not s3_bucket:
        raise HTTPException(
            status_code=500,
            detail="S3_OUTPUT_BUCKET environment variable is required",
        )

    s3 = get_s3_client()

    # Generate UUID for unique file name
    file_uuid = str(uuid.uuid4())

    # Construct S3 key: task_type/media/UUID.extension (no prefix)
    s3_key = f"{request.task_type}/media/{file_uuid}.{file_extension}"

    # Generate presigned URL (expires in 1 hour)
    expiration = 3600  # 1 hour in seconds

    presigned_url = s3.generate_presigned_url(
        "put_object",
        Params={
            "Bucket": s3_bucket,
            "Key": s3_key,
            "ContentType": request.content_type,
        },
        ExpiresIn=expiration,
    )

    return PresignedURLResponse(
        presigned_url=presigned_url,
        s3_path=f"s3://{s3_bucket}/{s3_key}",
        expires_in=expiration,
    )


def evaluate_provider(
    run_id: str,
    provider: str,
    language: str,
    input_dir: Path,
    output_dir: Path,
    port: int,
    s3_bucket: str,
) -> ProviderResult:
    """Evaluate a single STT provider."""
    try:
        s3 = get_s3_client()

        # Run pense STT eval command
        eval_cmd = [
            "pense",
            "stt",
            "eval",
            "-p",
            provider,
            "-l",
            language,
            "-i",
            str(input_dir),
            "-o",
            str(output_dir),
            "--port",
            str(port),
        ]

        print(f"Running {run_id} with command: ", " ".join(eval_cmd))

        subprocess.run(
            eval_cmd,
            capture_output=True,
            text=True,
            check=True,
            cwd=str(output_dir.parent),
        )

        # Find the provider-specific output directory
        provider_output_dir = None
        for item in output_dir.iterdir():
            if item.is_dir() and provider in item.name.lower():
                provider_output_dir = item
                break

        # Upload STT eval results to S3
        results_prefix = f"stt/evals/{run_id}/outputs/{provider}"

        # Read metrics.json and results.csv before uploading
        metrics_data = None
        results_data = None

        metrics_file = provider_output_dir / "metrics.json"
        results_file = provider_output_dir / "results.csv"

        if metrics_file.exists():
            with open(metrics_file, "r", encoding="utf-8") as f:
                metrics_data = json.load(f)

        if results_file.exists():
            results_data = []
            with open(results_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    results_data.append(dict(row))

        for root, dirs, files in os.walk(provider_output_dir):
            for file in files:
                local_file_path = Path(root) / file
                relative_path = local_file_path.relative_to(provider_output_dir)
                s3_key = f"{results_prefix}/{relative_path}"

                s3.upload_file(str(local_file_path), s3_bucket, s3_key)

        return ProviderResult(
            provider=provider,
            success=True,
            message=f"STT evaluation completed successfully for {provider}",
            metrics=metrics_data,
            results=results_data,
        )

    except subprocess.CalledProcessError as e:
        traceback.print_exc()
        return ProviderResult(
            provider=provider,
            success=False,
            message=f"STT eval failed: {e.stderr}",
        )
    except Exception as e:
        traceback.print_exc()
        return ProviderResult(
            provider=provider,
            success=False,
            message=f"Unexpected error: {str(e)}",
        )


def run_evaluation_task(
    task_id: str,
    request: STTEvaluationRequest,
    s3_bucket: str,
):
    """Run the STT evaluation in the background."""
    try:
        with tasks_lock:
            tasks[task_id]["status"] = TaskStatus.IN_PROGRESS.value

        s3 = get_s3_client()

        # Create temporary directory for processing
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)

            try:
                # Create directory structure
                input_dir = temp_path / "input"
                input_dir.mkdir()
                audios_wav_dir = input_dir / "audios" / "wav"
                audios_pcm16_dir = input_dir / "audios" / "pcm16"
                audios_wav_dir.mkdir(parents=True)
                audios_pcm16_dir.mkdir(parents=True)

                # Download audio files from S3 and create CSV
                stt_csv_path = input_dir / "stt.csv"
                with open(stt_csv_path, "w", newline="", encoding="utf-8") as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(["id", "text"])

                    for idx, (audio_path, gt_text) in enumerate(
                        zip(request.audio_paths, request.texts)
                    ):
                        # Parse S3 path (format: s3://bucket/key or bucket/key)
                        if audio_path.startswith("s3://"):
                            parts = audio_path[5:].split("/", 1)
                            bucket = parts[0]
                            key = parts[1] if len(parts) > 1 else ""
                        else:
                            parts = audio_path.split("/", 1)
                            bucket = parts[0]
                            key = parts[1] if len(parts) > 1 else ""

                        # Generate audio ID
                        audio_id = f"audio_{idx + 1}"

                        # Download audio file
                        local_wav_path = audios_wav_dir / f"{audio_id}.wav"
                        local_pcm16_path = audios_pcm16_dir / f"{audio_id}.wav"

                        s3.download_file(bucket, key, str(local_wav_path))
                        # Convert to pcm16 (mono, 16KHz, s16le) using ffmpeg instead of copying
                        cmd = [
                            "ffmpeg",
                            "-y",
                            "-i",
                            str(local_wav_path),
                            "-ac",
                            "1",
                            "-ar",
                            "16000",
                            "-f",
                            "wav",
                            "-sample_fmt",
                            "s16",
                            str(local_pcm16_path),
                        ]
                        subprocess.run(cmd, check=True)

                        # Write CSV row
                        writer.writerow([audio_id, gt_text])

                # Create output directory
                output_dir = temp_path / "output"
                output_dir.mkdir()

                # Find available ports for each provider
                provider_ports = {}
                start_port = 8000
                for provider in request.providers:
                    port = find_available_port(start_port)
                    provider_ports[provider] = port
                    start_port = port + 1

                # Run pense STT eval for all providers in parallel
                provider_results = []

                print(f"Running {len(request.providers)} providers in parallel")

                with concurrent.futures.ThreadPoolExecutor(
                    max_workers=len(request.providers)
                ) as executor:
                    future_to_provider = {
                        executor.submit(
                            evaluate_provider,
                            task_id,
                            provider,
                            request.language,
                            input_dir,
                            output_dir,
                            provider_ports[provider],
                            s3_bucket,
                        ): provider
                        for provider in request.providers
                    }

                    for future in concurrent.futures.as_completed(future_to_provider):
                        result = future.result()
                        provider_results.append(result)

                # Check if all providers succeeded
                all_succeeded = all(r.success for r in provider_results)
                if not all_succeeded:
                    failed_providers = [
                        r.provider for r in provider_results if not r.success
                    ]
                    with tasks_lock:
                        tasks[task_id]["status"] = TaskStatus.DONE.value
                        tasks[task_id]["provider_results"] = provider_results
                        tasks[task_id][
                            "error"
                        ] = f"Some providers failed: {', '.join(failed_providers)}"
                    return

                # Run pense STT leaderboard command
                leaderboard_dir = temp_path / "leaderboard"
                leaderboard_dir.mkdir()

                leaderboard_cmd = [
                    "pense",
                    "stt",
                    "leaderboard",
                    "-o",
                    str(output_dir),
                    "-s",
                    str(leaderboard_dir),
                ]

                leaderboard_prefix = f"stt/evals/{task_id}/leaderboard"
                leaderboard_summary = None
                metrics_path = None

                try:
                    subprocess.run(
                        leaderboard_cmd,
                        capture_output=True,
                        text=True,
                        check=True,
                        cwd=temp_path,
                    )

                    # Upload leaderboard results
                    for root, dirs, files in os.walk(leaderboard_dir):
                        for file in files:
                            local_file_path = Path(root) / file
                            relative_path = local_file_path.relative_to(leaderboard_dir)
                            s3_key = f"{leaderboard_prefix}/{relative_path}"

                            s3.upload_file(str(local_file_path), s3_bucket, s3_key)

                            # Read xlsx file and extract summary sheet
                            if file == "stt_leaderboard.xlsx" and openpyxl:
                                try:
                                    wb = openpyxl.load_workbook(
                                        str(local_file_path), data_only=True
                                    )
                                    if "summary" in wb.sheetnames:
                                        ws = wb["summary"]
                                        # Get headers from first row (skip empty cells)
                                        headers = [
                                            cell.value
                                            for cell in ws[1]
                                            if cell.value is not None
                                        ]
                                        # Read all data rows
                                        leaderboard_summary = []
                                        for row in ws.iter_rows(
                                            min_row=2, values_only=False
                                        ):
                                            # Check if row has any data
                                            if any(
                                                cell.value is not None for cell in row
                                            ):
                                                row_dict = {}
                                                for idx, cell in enumerate(row):
                                                    if idx < len(headers):
                                                        # Use header name as key, cell value as value
                                                        row_dict[headers[idx]] = (
                                                            cell.value
                                                        )
                                                # Only add row if it has at least one non-None value
                                                if any(
                                                    v is not None
                                                    for v in row_dict.values()
                                                ):
                                                    leaderboard_summary.append(row_dict)
                                except Exception as e:
                                    traceback.print_exc()
                                    # If reading xlsx fails, continue without summary
                                    pass

                            # Generate presigned URL for metrics image
                            if file == "all_metrics_by_run.png":
                                metrics_path = s3.generate_presigned_url(
                                    "get_object",
                                    Params={
                                        "Bucket": s3_bucket,
                                        "Key": s3_key,
                                    },
                                    ExpiresIn=3600,  # 1 hour
                                )

                except subprocess.CalledProcessError as e:
                    # Leaderboard failure is not critical, continue
                    pass

                # Update task with results
                with tasks_lock:
                    tasks[task_id]["status"] = TaskStatus.DONE.value
                    tasks[task_id]["provider_results"] = provider_results
                    tasks[task_id]["leaderboard_metrics_path"] = metrics_path
                    tasks[task_id]["leaderboard_summary"] = leaderboard_summary

            except Exception as e:
                traceback.print_exc()
                with tasks_lock:
                    tasks[task_id]["status"] = TaskStatus.DONE.value
                    tasks[task_id][
                        "error"
                    ] = f"Unexpected error during STT evaluation: {str(e)}"

    except Exception as e:
        traceback.print_exc()
        with tasks_lock:
            if task_id in tasks:
                tasks[task_id]["status"] = TaskStatus.DONE.value
                tasks[task_id]["error"] = f"Task failed: {str(e)}"


@app.post("/stt/evaluate", response_model=TaskCreateResponse)
async def evaluate_stt(request: STTEvaluationRequest):
    """
    Start a background task to evaluate multiple STT providers with audio files from S3.

    Returns a task ID that can be used to poll for status and results.
    """
    # Validate input
    if len(request.audio_paths) != len(request.texts):
        raise HTTPException(
            status_code=400,
            detail="Number of audio paths must match number of ground truth texts",
        )

    if not request.providers:
        raise HTTPException(
            status_code=400,
            detail="At least one provider must be specified",
        )

    # Get S3 configuration from environment
    try:
        s3_bucket = get_s3_output_config()
    except ValueError as e:
        raise HTTPException(status_code=500, detail=str(e))

    # Generate task ID
    task_id = str(uuid.uuid4())

    # Initialize task in storage
    with tasks_lock:
        tasks[task_id] = {
            "status": TaskStatus.IN_PROGRESS.value,
            "provider_results": None,
            "leaderboard_metrics_path": None,
            "leaderboard_summary": None,
            "error": None,
        }

    # Start background task in a separate thread
    thread = threading.Thread(
        target=run_evaluation_task,
        args=(task_id, request, s3_bucket),
        daemon=True,
    )
    thread.start()

    return TaskCreateResponse(task_id=task_id, status=TaskStatus.IN_PROGRESS.value)


@app.get("/stt/evaluate/{task_id}", response_model=TaskStatusResponse)
async def get_evaluation_status(task_id: str):
    """
    Get the status of an STT evaluation task.

    Returns the current status and, if done, the provider results and leaderboard path.
    """
    with tasks_lock:
        if task_id not in tasks:
            raise HTTPException(status_code=404, detail="Task not found")

        task = tasks[task_id]

    return TaskStatusResponse(
        task_id=task_id,
        status=task["status"],
        provider_results=task["provider_results"],
        leaderboard_metrics_path=task.get("leaderboard_metrics_path"),
        leaderboard_summary=task.get("leaderboard_summary"),
        error=task["error"],
    )
