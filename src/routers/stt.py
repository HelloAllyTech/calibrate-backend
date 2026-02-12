import os
import csv
import json
import subprocess
import tempfile
import time
import traceback
import threading
import logging
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
import openpyxl

from db import create_job, get_job, update_job
from auth_utils import get_current_user_id
from utils import (
    TaskStatus,
    ProviderResult,
    TaskCreateResponse,
    TaskStatusResponse,
    get_s3_client,
    get_s3_output_config,
    can_start_job,
    try_start_queued_job,
    register_job_starter,
    is_job_timed_out,
    kill_process_group,
    capture_exception_to_sentry,
)

# Job types that share the same queue
EVAL_JOB_TYPES = ["stt-eval", "tts-eval"]


def _start_stt_job_from_queue(job: dict) -> bool:
    """Start an STT evaluation job from the queue.

    This is called by the job queue manager when there's capacity to run a new job.
    """
    job_id = job["uuid"]
    details = job.get("details", {})

    # Reconstruct request from job details
    request = STTEvaluationRequest(
        audio_paths=details.get("audio_paths", []),
        texts=details.get("texts", []),
        providers=details.get("providers", []),
        language=details.get("language", ""),
    )
    s3_bucket = details.get("s3_bucket", "")

    # Start background task in a separate thread
    thread = threading.Thread(
        target=run_evaluation_task,
        args=(job_id, request, s3_bucket),
        daemon=True,
    )
    thread.start()

    return True


# Register the job starter for STT evaluation jobs
register_job_starter("stt-eval", _start_stt_job_from_queue)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/stt", tags=["stt"])


def _normalize_metrics(metrics):
    """Convert old list-of-dicts metrics format to new dict format.

    Old format: [{"wer": 2.4}, {"string_similarity": 0.15}, {"metric_name": "ttfb", "mean": 0.1, ...}, ...]
    New format: {"wer": 2.4, "string_similarity": 0.15, "ttfb": {"mean": 0.1, ...}, ...}
    """
    if metrics is None:
        return None
    if isinstance(metrics, dict):
        return metrics
    if isinstance(metrics, list):
        # Convert list of dicts to single dict
        result = {}
        for item in metrics:
            if isinstance(item, dict):
                # Check if it's a latency metric with metric_name field
                if "metric_name" in item:
                    metric_name = item["metric_name"]
                    # Create a copy without metric_name for the value
                    value = {k: v for k, v in item.items() if k != "metric_name"}
                    result[metric_name] = value
                else:
                    # Simple metric: {"wer": 2.4} - merge directly
                    result.update(item)
        return result if result else metrics  # Return original if conversion fails
    return metrics


def _collect_intermediate_results(output_dir: Path, providers: list) -> list:
    """Read whatever intermediate results are available from disk for each provider.

    Returns a list of ProviderResult objects preserving any partial results.
    """
    provider_results = []
    for provider in providers:
        provider_output_dir = _find_provider_output_dir(output_dir, provider)
        results_data = _read_results_csv(provider_output_dir)
        metrics_data = _read_metrics_json(provider_output_dir)
        if results_data:
            provider_results.append(
                ProviderResult(
                    provider=provider,
                    success=True,
                    metrics=metrics_data,
                    results=results_data,
                )
            )
        else:
            provider_results.append(
                ProviderResult(
                    provider=provider,
                    success=False,
                )
            )
    return provider_results


class STTEvaluationRequest(BaseModel):
    audio_paths: List[str]  # S3 paths to audio files
    texts: List[str]  # Ground truth text for each audio file
    providers: List[
        str
    ]  # List of STT providers (e.g., ["deepgram", "openai", "sarvam"])
    language: str  # Language (e.g., "english", "hindi")


def _find_provider_output_dir(output_dir: Path, provider: str) -> Optional[Path]:
    """Find the provider-specific output directory."""
    if not output_dir.exists():
        return None
    for item in output_dir.iterdir():
        if item.is_dir() and provider in item.name.lower():
            return item
    return None


def _read_results_csv(provider_output_dir: Path) -> Optional[List[dict]]:
    """Read results.csv from provider output directory if it exists."""
    if not provider_output_dir:
        return None
    results_file = provider_output_dir / "results.csv"
    if not results_file.exists():
        return None
    try:
        results_data = []
        with open(results_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                results_data.append(dict(row))
        return results_data
    except Exception:
        return None


def _read_metrics_json(provider_output_dir: Path) -> Optional[dict]:
    """Read metrics.json from provider output directory if it exists.

    Handles both new format (dict) and old format (list of dicts) for backward compatibility.
    """
    if not provider_output_dir:
        return None
    metrics_file = provider_output_dir / "metrics.json"
    if not metrics_file.exists():
        return None
    try:
        with open(metrics_file, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data
    except Exception:
        return None


def _read_leaderboard_xlsx(leaderboard_dir: Path) -> Optional[List[dict]]:
    """Read the leaderboard summary from the xlsx file in leaderboard directory.

    Looks for any .xlsx file in the directory (commonly stt_leaderboard.xlsx).
    """
    if not leaderboard_dir.exists():
        logger.warning(f"Leaderboard directory does not exist: {leaderboard_dir}")
        return None

    # Find xlsx file in leaderboard directory
    xlsx_files = list(leaderboard_dir.glob("*.xlsx"))
    if not xlsx_files:
        logger.warning(
            f"No xlsx files found in leaderboard directory: {leaderboard_dir}"
        )
        # Log what files are present for debugging
        all_files = list(leaderboard_dir.iterdir())
        logger.info(f"Files in leaderboard directory: {[f.name for f in all_files]}")
        return None

    xlsx_file = xlsx_files[0]  # Use the first xlsx file found
    logger.info(f"Reading leaderboard from: {xlsx_file}")

    try:
        wb = openpyxl.load_workbook(str(xlsx_file), data_only=True)
        logger.info(f"Workbook sheets: {wb.sheetnames}")

        if "summary" not in wb.sheetnames:
            logger.warning(
                f"'summary' sheet not found in {xlsx_file.name}, sheets: {wb.sheetnames}"
            )
            return None

        ws = wb["summary"]
        # Get headers from first row (skip empty cells)
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        logger.info(f"Leaderboard headers: {headers}")

        leaderboard_summary = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            if any(cell.value is not None for cell in row):
                row_dict = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = cell.value
                if any(v is not None for v in row_dict.values()):
                    leaderboard_summary.append(row_dict)

        logger.info(f"Read {len(leaderboard_summary)} rows from leaderboard")
        return leaderboard_summary
    except Exception as e:
        logger.warning(f"Failed to read leaderboard xlsx: {e}")
        return None


def run_evaluation_task(
    task_id: str,
    request: STTEvaluationRequest,
    s3_bucket: str,
):
    """Run the STT evaluation in the background."""
    try:
        logger.info(
            f"Running evaluation task {task_id} with {len(request.providers)} providers"
        )
        update_job(task_id, status=TaskStatus.IN_PROGRESS.value)

        s3 = get_s3_client()

        # Create temporary directory for processing
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)

            try:
                # Create directory structure
                input_dir = temp_path / "input"
                input_dir.mkdir()
                audios_dir = input_dir / "audios"
                audios_dir.mkdir(parents=True)

                # Download audio files from S3 and create CSV
                stt_csv_path = input_dir / "stt.csv"
                with open(stt_csv_path, "w", newline="", encoding="utf-8") as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(["id", "text"])

                    for idx, (audio_path, gt_text) in enumerate(
                        zip(request.audio_paths, request.texts)
                    ):
                        # Parse S3 path (format: s3://bucket/key or bucket/key)
                        if audio_path.startswith("s3://"):
                            parts = audio_path[5:].split("/", 1)
                            bucket = parts[0]
                            key = parts[1] if len(parts) > 1 else ""
                        else:
                            parts = audio_path.split("/", 1)
                            bucket = parts[0]
                            key = parts[1] if len(parts) > 1 else ""

                        # Generate audio ID
                        audio_id = f"audio_{idx + 1}"

                        # Download audio file directly to audios folder
                        local_audio_path = audios_dir / f"{audio_id}.wav"

                        logger.info(
                            f"Downloading audio file from {bucket}/{key} to {local_audio_path}"
                        )
                        s3.download_file(bucket, key, str(local_audio_path))

                        # Write CSV row
                        writer.writerow([audio_id, gt_text])

                # Create output directory
                output_dir = temp_path / "output"
                output_dir.mkdir()

                # Run calibrate stt command with all providers at once
                # The CLI now handles parallelization internally and generates leaderboard
                eval_cmd = (
                    [
                        "calibrate",
                        "stt",
                        "-p",
                    ]
                    + request.providers
                    + [
                        "-l",
                        request.language,
                        "-i",
                        str(input_dir),
                        "-o",
                        str(output_dir),
                    ]
                )

                logger.info(f"Running STT eval command: {' '.join(eval_cmd)}")

                # Create temp files for stdout/stderr
                stdout_path = output_dir / "stdout.log"
                stderr_path = output_dir / "stderr.log"

                with (
                    open(stdout_path, "w") as stdout_f,
                    open(stderr_path, "w") as stderr_f,
                ):
                    process = subprocess.Popen(
                        eval_cmd,
                        stdout=stdout_f,
                        stderr=stderr_f,
                        text=True,
                        start_new_session=True,
                        cwd=str(temp_path),
                    )

                    # Store PID and output dir for cleanup and intermediate results
                    update_job(
                        task_id,
                        details={
                            "pid": process.pid,
                            "pgid": process.pid,
                            "output_dir": str(output_dir),
                        },
                    )

                    # Poll for process completion with heartbeat to keep updated_at fresh
                    # This prevents the job from being marked as timed out during long runs
                    HEARTBEAT_INTERVAL = 2  # seconds
                    while process.poll() is None:
                        time.sleep(HEARTBEAT_INTERVAL)
                        if process.poll() is None:
                            # Process still running, send heartbeat to refresh updated_at
                            update_job(task_id)

                # Read stdout/stderr
                with open(stdout_path, "r") as f:
                    stdout = f.read()
                with open(stderr_path, "r") as f:
                    stderr = f.read()

                if process.returncode != 0:
                    logger.error(f"STT eval failed with code {process.returncode}")
                    logger.error(f"stderr: {stderr}")
                    raise subprocess.CalledProcessError(
                        process.returncode, eval_cmd, stdout, stderr
                    )

                logger.info("STT eval command completed successfully")

                # Read results for each provider
                provider_results = []
                for provider in request.providers:
                    provider_output_dir = _find_provider_output_dir(
                        output_dir, provider
                    )
                    if provider_output_dir:
                        metrics_data = _read_metrics_json(provider_output_dir)
                        results_data = _read_results_csv(provider_output_dir)

                        # Upload provider results to S3
                        results_prefix = f"stt/evals/{task_id}/outputs/{provider}"
                        for root, dirs, files in os.walk(provider_output_dir):
                            for file in files:
                                local_file_path = Path(root) / file
                                relative_path = local_file_path.relative_to(
                                    provider_output_dir
                                )
                                s3_key = f"{results_prefix}/{relative_path}"
                                s3.upload_file(str(local_file_path), s3_bucket, s3_key)

                        provider_results.append(
                            ProviderResult(
                                provider=provider,
                                success=True,
                                metrics=metrics_data,
                                results=results_data,
                            )
                        )
                    else:
                        provider_results.append(
                            ProviderResult(
                                provider=provider,
                                success=False,
                            )
                        )

                # Read leaderboard from output directory
                leaderboard_dir = output_dir / "leaderboard"
                leaderboard_summary = None

                # Log what's in output_dir for debugging
                logger.info(
                    f"Output directory contents: {[f.name for f in output_dir.iterdir()]}"
                )

                if leaderboard_dir.exists():
                    logger.info(f"Leaderboard directory exists: {leaderboard_dir}")
                    leaderboard_summary = _read_leaderboard_xlsx(leaderboard_dir)

                    # Upload leaderboard to S3
                    leaderboard_prefix = f"stt/evals/{task_id}/leaderboard"
                    for root, dirs, files in os.walk(leaderboard_dir):
                        for file in files:
                            local_file_path = Path(root) / file
                            relative_path = local_file_path.relative_to(leaderboard_dir)
                            s3_key = f"{leaderboard_prefix}/{relative_path}"
                            s3.upload_file(str(local_file_path), s3_bucket, s3_key)
                else:
                    logger.warning(
                        f"Leaderboard directory does not exist: {leaderboard_dir}"
                    )

                # Create and upload config file to S3
                config_data = {
                    "providers": request.providers,
                    "language": request.language,
                    "audio_count": len(request.audio_paths),
                }
                config_file = temp_path / "config.json"
                with open(config_file, "w", encoding="utf-8") as f:
                    json.dump(config_data, f, indent=2)
                config_s3_key = f"stt/evals/{task_id}/config.json"
                s3.upload_file(str(config_file), s3_bucket, config_s3_key)
                logger.info(f"Uploaded config file to S3: {config_s3_key}")

                # Check if all providers succeeded
                all_succeeded = all(r.success for r in provider_results)
                final_status = (
                    TaskStatus.DONE.value if all_succeeded else TaskStatus.FAILED.value
                )

                error_msg = None
                if not all_succeeded:
                    failed = [r.provider for r in provider_results if not r.success]
                    error_msg = f"Some providers failed: {', '.join(failed)}"

                # Update job with results
                update_job(
                    task_id,
                    status=final_status,
                    results={
                        "provider_results": [r.model_dump() for r in provider_results],
                        "leaderboard_summary": leaderboard_summary,
                        "error": error_msg,
                    },
                )

            except subprocess.CalledProcessError as e:
                traceback.print_exc()
                capture_exception_to_sentry(e)
                error_results = {
                    "error": f"STT evaluation failed: {e.stderr if hasattr(e, 'stderr') else str(e)}",
                }
                # Preserve any intermediate results already written to disk
                try:
                    if output_dir.exists():
                        intermediate = _collect_intermediate_results(
                            output_dir, request.providers
                        )
                        if intermediate:
                            error_results["provider_results"] = [
                                r.model_dump() for r in intermediate
                            ]
                except Exception:
                    pass
                update_job(
                    task_id,
                    status=TaskStatus.FAILED.value,
                    results=error_results,
                )
            except Exception as e:
                traceback.print_exc()
                capture_exception_to_sentry(e)
                error_results = {
                    "error": f"Unexpected error during STT evaluation: {str(e)}",
                }
                # Preserve any intermediate results already written to disk
                try:
                    if output_dir.exists():
                        intermediate = _collect_intermediate_results(
                            output_dir, request.providers
                        )
                        if intermediate:
                            error_results["provider_results"] = [
                                r.model_dump() for r in intermediate
                            ]
                except Exception:
                    pass
                update_job(
                    task_id,
                    status=TaskStatus.FAILED.value,
                    results=error_results,
                )

    except Exception as e:
        traceback.print_exc()
        capture_exception_to_sentry(e)
        update_job(
            task_id,
            status=TaskStatus.FAILED.value,
            results={"error": f"Task failed: {str(e)}"},
        )
    finally:
        # Try to start the next queued job
        try_start_queued_job(EVAL_JOB_TYPES)


@router.post("/evaluate", response_model=TaskCreateResponse)
async def evaluate_stt(
    request: STTEvaluationRequest, user_id: str = Depends(get_current_user_id)
):
    """
    Start a background task to evaluate multiple STT providers with audio files from S3.

    Returns a task ID that can be used to poll for status and results.
    """
    # Validate input
    if len(request.audio_paths) != len(request.texts):
        raise HTTPException(
            status_code=400,
            detail="Number of audio paths must match number of ground truth texts",
        )

    if not request.providers:
        raise HTTPException(
            status_code=400,
            detail="At least one provider must be specified",
        )

    # Get S3 configuration from environment
    try:
        s3_bucket = get_s3_output_config()
    except ValueError as e:
        raise HTTPException(status_code=500, detail=str(e))

    # Check if we can start immediately or need to queue
    can_start = can_start_job(EVAL_JOB_TYPES, user_id)
    initial_status = (
        TaskStatus.IN_PROGRESS.value if can_start else TaskStatus.QUEUED.value
    )

    # Create job in database with details for recovery
    job_id = create_job(
        job_type="stt-eval",
        user_id=user_id,
        status=initial_status,
        details={
            "audio_paths": request.audio_paths,
            "texts": request.texts,
            "providers": request.providers,
            "language": request.language,
            "s3_bucket": s3_bucket,
        },
        results=None,
    )

    if can_start:
        # Start background task in a separate thread
        thread = threading.Thread(
            target=run_evaluation_task,
            args=(job_id, request, s3_bucket),
            daemon=True,
        )
        thread.start()
        logger.info(f"Started STT evaluation job {job_id} immediately")
    else:
        logger.info(f"Queued STT evaluation job {job_id}")

    return TaskCreateResponse(task_id=job_id, status=initial_status)


@router.get("/evaluate/{task_id}", response_model=TaskStatusResponse)
async def get_evaluation_status(
    task_id: str, user_id: str = Depends(get_current_user_id)
):
    """
    Get the status of an STT evaluation task.

    Returns the current status and, if done, the provider results and leaderboard path.
    """
    job = get_job(task_id, user_id=user_id)
    if not job:
        raise HTTPException(status_code=404, detail="Task not found")

    status = job["status"]
    results = job.get("results") or {}
    details = job.get("details") or {}

    # Check for timeout on in-progress jobs
    if status == TaskStatus.IN_PROGRESS.value:
        updated_at = job.get("updated_at")
        if updated_at and is_job_timed_out(updated_at):
            logger.warning(f"Job {task_id} timed out, marking as failed")

            # Kill running process
            pid = details.get("pid") or details.get("pgid")
            if pid:
                kill_process_group(pid, task_id)

            # Preserve intermediate results from disk before marking as failed
            # IMPORTANT: Merge with existing results, don't overwrite successful ones
            requested_providers = details.get("providers", [])
            output_dir_str = details.get("output_dir")
            existing_provider_results = results.get("provider_results", [])

            # Build a map of existing successful results (don't overwrite these)
            existing_success_map = {}
            for pr in existing_provider_results:
                if pr.get("success") is True:
                    existing_success_map[pr.get("provider")] = pr

            if output_dir_str:
                try:
                    output_dir = Path(output_dir_str)
                    if output_dir.exists():
                        intermediate = _collect_intermediate_results(
                            output_dir, requested_providers
                        )
                        # Merge: keep existing successful results, add new ones from disk
                        merged_results = []
                        intermediate_map = {
                            r.provider: r.model_dump() for r in intermediate
                        }
                        for provider in requested_providers:
                            if provider in existing_success_map:
                                # Keep the existing successful result
                                merged_results.append(existing_success_map[provider])
                            elif provider in intermediate_map:
                                # Use intermediate result from disk
                                merged_results.append(intermediate_map[provider])
                            else:
                                # Provider not found anywhere, mark as failed
                                merged_results.append(
                                    {
                                        "provider": provider,
                                        "success": False,
                                        "metrics": None,
                                        "results": None,
                                    }
                                )
                        results["provider_results"] = merged_results
                except Exception as exc:
                    logger.warning(
                        f"Failed to collect intermediate results on timeout: {exc}"
                    )
                    # Even on exception, preserve any existing successful results
                    if existing_provider_results:
                        results["provider_results"] = existing_provider_results

            # Mark job as failed
            results["error"] = "Job timed out after 5 minutes of inactivity"
            update_job(
                task_id,
                status=TaskStatus.FAILED.value,
                results=results,
            )
            status = TaskStatus.FAILED.value

            # Try to start the next queued job
            try_start_queued_job(EVAL_JOB_TYPES)

    # Get list of all requested providers from job details
    requested_providers = details.get("providers", [])

    # Build provider results
    provider_results = results.get("provider_results")
    if provider_results is None and status == TaskStatus.IN_PROGRESS.value:
        # Job is in progress - try to read intermediate results from disk
        output_dir_str = details.get("output_dir")
        expected_total = len(details.get("audio_paths", []))
        if output_dir_str:
            output_dir = Path(output_dir_str)
            provider_results = []
            for provider in requested_providers:
                provider_output_dir = _find_provider_output_dir(output_dir, provider)
                results_data = _read_results_csv(provider_output_dir)
                metrics_data = _read_metrics_json(provider_output_dir)
                if results_data:
                    # If all files are processed and metrics are ready, mark as done
                    provider_done = (
                        len(results_data) >= expected_total and metrics_data is not None
                    )
                    provider_results.append(
                        {
                            "provider": provider,
                            "success": True if provider_done else None,
                            "message": (
                                f"Done ({len(results_data)} files processed)"
                                if provider_done
                                else f"Running... ({len(results_data)} files processed)"
                            ),
                            "metrics": metrics_data,
                            "results": results_data,
                        }
                    )
                else:
                    provider_results.append(
                        {
                            "provider": provider,
                            "success": None,
                            "message": "Queued...",
                            "metrics": None,
                            "results": None,
                        }
                    )

    if provider_results is None:
        # Job hasn't completed yet or no output dir available, show all as queued
        provider_results = [
            {
                "provider": provider,
                "success": None,
                "message": "Queued...",
                "metrics": None,
                "results": None,
            }
            for provider in requested_providers
        ]

    # Normalize metrics format for backward compatibility (list -> dict)
    for provider_result in provider_results:
        if provider_result.get("metrics"):
            provider_result["metrics"] = _normalize_metrics(provider_result["metrics"])

    return TaskStatusResponse(
        task_id=task_id,
        status=status,
        language=details.get("language"),
        provider_results=provider_results,
        leaderboard_summary=results.get("leaderboard_summary"),
        error=results.get("error"),
    )
