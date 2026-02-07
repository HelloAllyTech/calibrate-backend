import os
import csv
import json
import subprocess
import tempfile
import traceback
import threading
import logging
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
import openpyxl

from db import create_job, get_job, update_job
from auth_utils import get_current_user_id
from utils import (
    TaskStatus,
    ProviderResult,
    TaskCreateResponse,
    TaskStatusResponse,
    get_s3_client,
    get_s3_output_config,
    can_start_job,
    try_start_queued_job,
    register_job_starter,
    generate_presigned_download_url,
    is_job_timed_out,
    kill_process_group,
    capture_exception_to_sentry,
)

# Job types that share the same queue
EVAL_JOB_TYPES = ["stt-eval", "tts-eval"]


def _start_tts_job_from_queue(job: dict) -> bool:
    """Start a TTS evaluation job from the queue.

    This is called by the job queue manager when there's capacity to run a new job.
    """
    job_id = job["uuid"]
    details = job.get("details", {})

    # Reconstruct request from job details
    request = TTSEvaluationRequest(
        texts=details.get("texts", []),
        providers=details.get("providers", []),
        language=details.get("language", ""),
    )
    s3_bucket = details.get("s3_bucket", "")

    # Start background task in a separate thread
    thread = threading.Thread(
        target=run_tts_evaluation_task,
        args=(job_id, request, s3_bucket),
        daemon=True,
    )
    thread.start()

    return True


# Register the job starter for TTS evaluation jobs
register_job_starter("tts-eval", _start_tts_job_from_queue)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/tts", tags=["tts"])


def _normalize_metrics(metrics):
    """Convert old list-of-dicts metrics format to new dict format.

    Old format: [{"wer": 2.4}, {"string_similarity": 0.15}, {"metric_name": "ttfb", "mean": 0.1, ...}, ...]
    New format: {"wer": 2.4, "string_similarity": 0.15, "ttfb": {"mean": 0.1, ...}, ...}
    """
    if metrics is None:
        return None
    if isinstance(metrics, dict):
        return metrics
    if isinstance(metrics, list):
        # Convert list of dicts to single dict
        result = {}
        for item in metrics:
            if isinstance(item, dict):
                # Check if it's a latency metric with metric_name field
                if "metric_name" in item:
                    metric_name = item["metric_name"]
                    # Create a copy without metric_name for the value
                    value = {k: v for k, v in item.items() if k != "metric_name"}
                    result[metric_name] = value
                else:
                    # Simple metric: {"wer": 2.4} - merge directly
                    result.update(item)
        return result if result else metrics  # Return original if conversion fails
    return metrics


class TTSEvaluationRequest(BaseModel):
    texts: List[str]  # List of texts to synthesize
    providers: List[
        str
    ]  # List of TTS providers (e.g., ["smallest", "cartesia", "openai"])
    language: str  # Language (e.g., "english", "hindi")


def _find_tts_provider_output_dir(output_dir: Path, provider: str) -> Optional[Path]:
    """Find the provider-specific output directory."""
    if not output_dir.exists():
        return None
    for item in output_dir.iterdir():
        if item.is_dir() and provider in item.name.lower():
            return item
    return None


def _read_tts_results_csv(provider_output_dir: Path) -> Optional[List[dict]]:
    """Read results.csv from provider output directory if it exists."""
    if not provider_output_dir:
        return None
    results_file = provider_output_dir / "results.csv"
    if not results_file.exists():
        return None
    try:
        results_data = []
        with open(results_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                results_data.append(dict(row))
        return results_data
    except Exception:
        return None


def _read_tts_metrics_json(provider_output_dir: Path) -> Optional[dict]:
    """Read metrics.json from provider output directory if it exists.

    Handles both new format (dict) and old format (list of dicts) for backward compatibility.
    """
    if not provider_output_dir:
        return None
    metrics_file = provider_output_dir / "metrics.json"
    if not metrics_file.exists():
        return None
    try:
        with open(metrics_file, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data
    except Exception:
        return None


def _read_leaderboard_xlsx(leaderboard_dir: Path) -> Optional[List[dict]]:
    """Read the leaderboard summary from the xlsx file in leaderboard directory.

    Looks for any .xlsx file in the directory (commonly tts_leaderboard.xlsx).
    """
    if not leaderboard_dir.exists():
        logger.warning(f"Leaderboard directory does not exist: {leaderboard_dir}")
        return None

    # Find xlsx file in leaderboard directory
    xlsx_files = list(leaderboard_dir.glob("*.xlsx"))
    if not xlsx_files:
        logger.warning(
            f"No xlsx files found in leaderboard directory: {leaderboard_dir}"
        )
        # Log what files are present for debugging
        all_files = list(leaderboard_dir.iterdir())
        logger.info(f"Files in leaderboard directory: {[f.name for f in all_files]}")
        return None

    xlsx_file = xlsx_files[0]  # Use the first xlsx file found
    logger.info(f"Reading leaderboard from: {xlsx_file}")

    try:
        wb = openpyxl.load_workbook(str(xlsx_file), data_only=True)
        logger.info(f"Workbook sheets: {wb.sheetnames}")

        if "summary" not in wb.sheetnames:
            logger.warning(
                f"'summary' sheet not found in {xlsx_file.name}, sheets: {wb.sheetnames}"
            )
            return None

        ws = wb["summary"]
        # Get headers from first row (skip empty cells)
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        logger.info(f"Leaderboard headers: {headers}")

        leaderboard_summary = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            if any(cell.value is not None for cell in row):
                row_dict = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = cell.value
                if any(v is not None for v in row_dict.values()):
                    leaderboard_summary.append(row_dict)

        logger.info(f"Read {len(leaderboard_summary)} rows from leaderboard")
        return leaderboard_summary
    except Exception as e:
        logger.warning(f"Failed to read leaderboard xlsx: {e}")
        return None


def run_tts_evaluation_task(
    task_id: str,
    request: TTSEvaluationRequest,
    s3_bucket: str,
):
    """Run the TTS evaluation in the background."""
    try:
        logger.info(
            f"Running TTS evaluation task {task_id} with {len(request.providers)} providers"
        )
        update_job(task_id, status=TaskStatus.IN_PROGRESS.value)

        s3 = get_s3_client()

        # Create temporary directory for processing
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)

            try:
                # Create input CSV file
                input_csv = temp_path / "input.csv"
                with open(input_csv, "w", newline="", encoding="utf-8") as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(["id", "text"])
                    for idx, text in enumerate(request.texts):
                        writer.writerow([idx, text])

                # Create output directory
                output_dir = temp_path / "output"
                output_dir.mkdir()

                # Run calibrate tts command with all providers at once
                # The CLI now handles parallelization internally and generates leaderboard
                eval_cmd = (
                    [
                        "calibrate",
                        "tts",
                        "-p",
                    ]
                    + request.providers
                    + [
                        "-l",
                        request.language,
                        "-i",
                        str(input_csv),
                        "-o",
                        str(output_dir),
                    ]
                )

                logger.info(f"Running TTS eval command: {' '.join(eval_cmd)}")

                # Create temp files for stdout/stderr
                stdout_path = output_dir / "stdout.log"
                stderr_path = output_dir / "stderr.log"

                with (
                    open(stdout_path, "w") as stdout_f,
                    open(stderr_path, "w") as stderr_f,
                ):
                    process = subprocess.Popen(
                        eval_cmd,
                        stdout=stdout_f,
                        stderr=stderr_f,
                        text=True,
                        start_new_session=True,
                        cwd=str(temp_path),
                    )

                    # Store PID for cleanup
                    update_job(
                        task_id, details={"pid": process.pid, "pgid": process.pid}
                    )

                    # Wait for process to complete
                    process.wait()

                # Read stdout/stderr
                with open(stdout_path, "r") as f:
                    stdout = f.read()
                with open(stderr_path, "r") as f:
                    stderr = f.read()

                if process.returncode != 0:
                    logger.error(f"TTS eval failed with code {process.returncode}")
                    logger.error(f"stderr: {stderr}")
                    raise subprocess.CalledProcessError(
                        process.returncode, eval_cmd, stdout, stderr
                    )

                logger.info("TTS eval command completed successfully")

                # Read results for each provider
                provider_results = []
                for provider in request.providers:
                    provider_output_dir = _find_tts_provider_output_dir(
                        output_dir, provider
                    )
                    if provider_output_dir:
                        metrics_data = _read_tts_metrics_json(provider_output_dir)
                        results_data = _read_tts_results_csv(provider_output_dir)

                        # Upload provider results to S3 and map audio paths
                        results_prefix = f"tts/evals/{task_id}/outputs/{provider}"
                        audio_path_to_s3_key = {}

                        for root, dirs, files in os.walk(provider_output_dir):
                            for file in files:
                                local_file_path = Path(root) / file
                                relative_path = local_file_path.relative_to(
                                    provider_output_dir
                                )
                                s3_key = f"{results_prefix}/{relative_path}"
                                s3.upload_file(str(local_file_path), s3_bucket, s3_key)

                                # Track audio files for path mapping
                                if file.endswith((".wav", ".mp3", ".ogg")):
                                    audio_path_to_s3_key[str(local_file_path)] = s3_key

                        # Replace local audio paths with S3 keys in results
                        successful_count = 0
                        if results_data:
                            for result_row in results_data:
                                if (
                                    "audio_path" in result_row
                                    and result_row["audio_path"]
                                ):
                                    local_audio_path = result_row["audio_path"]
                                    audio_s3_key = audio_path_to_s3_key.get(
                                        local_audio_path
                                    )
                                    if audio_s3_key:
                                        result_row["audio_path"] = audio_s3_key
                                        successful_count += 1

                        if successful_count > 0:
                            provider_results.append(
                                ProviderResult(
                                    provider=provider,
                                    success=True,
                                    message=f"TTS evaluation completed successfully for {provider}",
                                    metrics=metrics_data,
                                    results=results_data,
                                )
                            )
                        else:
                            provider_results.append(
                                ProviderResult(
                                    provider=provider,
                                    success=False,
                                    message=f"TTS evaluation completed with errors for {provider}: no texts synthesized successfully",
                                    metrics=metrics_data,
                                    results=results_data,
                                )
                            )
                    else:
                        provider_results.append(
                            ProviderResult(
                                provider=provider,
                                success=False,
                                message=f"No output found for provider {provider}",
                            )
                        )

                # Read leaderboard from output directory
                leaderboard_dir = output_dir / "leaderboard"
                leaderboard_summary = None

                # Log what's in output_dir for debugging
                logger.info(
                    f"Output directory contents: {[f.name for f in output_dir.iterdir()]}"
                )

                if leaderboard_dir.exists():
                    logger.info(f"Leaderboard directory exists: {leaderboard_dir}")
                    leaderboard_summary = _read_leaderboard_xlsx(leaderboard_dir)

                    # Upload leaderboard to S3
                    leaderboard_prefix = f"tts/evals/{task_id}/leaderboard"
                    for root, dirs, files in os.walk(leaderboard_dir):
                        for file in files:
                            local_file_path = Path(root) / file
                            relative_path = local_file_path.relative_to(leaderboard_dir)
                            s3_key = f"{leaderboard_prefix}/{relative_path}"
                            s3.upload_file(str(local_file_path), s3_bucket, s3_key)
                else:
                    logger.warning(
                        f"Leaderboard directory does not exist: {leaderboard_dir}"
                    )

                # Create and upload config file to S3
                config_data = {
                    "providers": request.providers,
                    "language": request.language,
                    "text_count": len(request.texts),
                }
                config_file = temp_path / "config.json"
                with open(config_file, "w", encoding="utf-8") as f:
                    json.dump(config_data, f, indent=2)
                config_s3_key = f"tts/evals/{task_id}/config.json"
                s3.upload_file(str(config_file), s3_bucket, config_s3_key)
                logger.info(f"Uploaded config file to S3: {config_s3_key}")

                # Check if all providers succeeded
                all_succeeded = all(r.success for r in provider_results)
                final_status = (
                    TaskStatus.DONE.value if all_succeeded else TaskStatus.FAILED.value
                )

                error_msg = None
                if not all_succeeded:
                    failed = [r.provider for r in provider_results if not r.success]
                    error_msg = f"Some providers failed: {', '.join(failed)}"

                # Update job with results
                update_job(
                    task_id,
                    status=final_status,
                    results={
                        "provider_results": [r.model_dump() for r in provider_results],
                        "leaderboard_summary": leaderboard_summary,
                        "error": error_msg,
                    },
                )

            except subprocess.CalledProcessError as e:
                traceback.print_exc()
                capture_exception_to_sentry(e)
                update_job(
                    task_id,
                    status=TaskStatus.FAILED.value,
                    results={
                        "error": f"TTS evaluation failed: {e.stderr if hasattr(e, 'stderr') else str(e)}",
                    },
                )
            except Exception as e:
                traceback.print_exc()
                capture_exception_to_sentry(e)
                update_job(
                    task_id,
                    status=TaskStatus.FAILED.value,
                    results={
                        "error": f"Unexpected error during TTS evaluation: {str(e)}",
                    },
                )

    except Exception as e:
        traceback.print_exc()
        capture_exception_to_sentry(e)
        update_job(
            task_id,
            status=TaskStatus.FAILED.value,
            results={"error": f"Task failed: {str(e)}"},
        )
    finally:
        # Try to start the next queued job
        try_start_queued_job(EVAL_JOB_TYPES)


@router.post("/evaluate", response_model=TaskCreateResponse)
async def evaluate_tts(
    request: TTSEvaluationRequest, user_id: str = Depends(get_current_user_id)
):
    """
    Start a background task to evaluate multiple TTS providers with text inputs.

    Returns a task ID that can be used to poll for status and results.
    """
    # Validate input
    if not request.texts:
        raise HTTPException(
            status_code=400,
            detail="At least one text must be provided",
        )

    if not request.providers:
        raise HTTPException(
            status_code=400,
            detail="At least one provider must be specified",
        )

    # Get S3 configuration from environment
    try:
        s3_bucket = get_s3_output_config()
    except ValueError as e:
        raise HTTPException(status_code=500, detail=str(e))

    # Check if we can start immediately or need to queue
    can_start = can_start_job(EVAL_JOB_TYPES)
    initial_status = (
        TaskStatus.IN_PROGRESS.value if can_start else TaskStatus.QUEUED.value
    )

    # Create job in database with details for recovery
    job_id = create_job(
        job_type="tts-eval",
        user_id=user_id,
        status=initial_status,
        details={
            "texts": request.texts,
            "providers": request.providers,
            "language": request.language,
            "s3_bucket": s3_bucket,
        },
        results=None,
    )

    if can_start:
        # Start background task in a separate thread
        thread = threading.Thread(
            target=run_tts_evaluation_task,
            args=(job_id, request, s3_bucket),
            daemon=True,
        )
        thread.start()
        logger.info(f"Started TTS evaluation job {job_id} immediately")
    else:
        logger.info(f"Queued TTS evaluation job {job_id}")

    return TaskCreateResponse(task_id=job_id, status=initial_status)


@router.get("/evaluate/{task_id}", response_model=TaskStatusResponse)
async def get_tts_evaluation_status(
    task_id: str, user_id: str = Depends(get_current_user_id)
):
    """
    Get the status of a TTS evaluation task.

    Returns the current status and, if done, the provider results and leaderboard path.
    """
    job = get_job(task_id, user_id=user_id)
    if not job:
        raise HTTPException(status_code=404, detail="Task not found")

    status = job["status"]
    results = job.get("results") or {}
    details = job.get("details") or {}

    # Check for timeout on in-progress jobs
    if status == TaskStatus.IN_PROGRESS.value:
        updated_at = job.get("updated_at")
        if updated_at and is_job_timed_out(updated_at):
            logger.warning(f"Job {task_id} timed out, marking as failed")

            # Kill running process
            pid = details.get("pid") or details.get("pgid")
            if pid:
                kill_process_group(pid, task_id)

            # Mark job as failed
            results["error"] = "Job timed out after 5 minutes of inactivity"
            update_job(
                task_id,
                status=TaskStatus.FAILED.value,
                results=results,
            )
            status = TaskStatus.FAILED.value

            # Try to start the next queued job
            try_start_queued_job(EVAL_JOB_TYPES)

    # Get list of all requested providers from job details
    requested_providers = details.get("providers", [])

    # Build provider results
    provider_results = results.get("provider_results")
    if provider_results is None:
        # Job hasn't completed yet, show all as queued
        provider_results = [
            {
                "provider": provider,
                "success": None,
                "message": "Queued...",
                "metrics": None,
                "results": None,
            }
            for provider in requested_providers
        ]

    # Normalize metrics format for backward compatibility (list -> dict)
    for provider_result in provider_results:
        if provider_result.get("metrics"):
            provider_result["metrics"] = _normalize_metrics(provider_result["metrics"])

    # Generate presigned URLs on the fly for completed jobs
    if status == TaskStatus.DONE.value:
        for provider_result in provider_results:
            if provider_result.get("results"):
                for result_row in provider_result["results"]:
                    if "audio_path" in result_row and result_row["audio_path"]:
                        audio_s3_key = result_row["audio_path"]
                        # Skip if already a URL (backwards compatibility)
                        if audio_s3_key.startswith("http") or audio_s3_key.startswith(
                            "s3://"
                        ):
                            continue
                        presigned_url = generate_presigned_download_url(audio_s3_key)
                        if presigned_url:
                            result_row["audio_path"] = presigned_url

    return TaskStatusResponse(
        task_id=task_id,
        status=status,
        language=details.get("language"),
        provider_results=provider_results,
        leaderboard_summary=results.get("leaderboard_summary"),
        error=results.get("error"),
    )
